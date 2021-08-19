import openpyxl
import requests

class  Excel_read:
    def getExcel(self, path):

        wb = openpyxl.load_workbook(path)
        sheet =wb['Sheet1']
        cases = sheet.max_row
        list2 =[]
        for case in range(2,cases+1):
            dict_case = dict(method=sheet.cell(row=case,column=1).value,
                             url= sheet.cell(row=case,column=2).value,
                             body=sheet.cell(row=case, column=3).value,)
            list2.append(dict_case)
        return list2


